import openpyxl
from openpyxl.styles import Font, Alignment
from django.contrib import admin
from .models import Equipamento, Movimentacao
from django.http import HttpResponse
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet


# Filtro customizado para patrim√¥nio
class PatrimonioFilter(admin.SimpleListFilter):
    title = 'Possui patrim√¥nio'
    parameter_name = 'patrimonio'

    def lookups(self, request, model_admin):
        return (
            ('com', 'Com patrim√¥nio'),
            ('sem', 'Sem patrim√¥nio'),
        )

    def queryset(self, request, queryset):
        if self.value() == 'com':
            return queryset.exclude(numero_patrimonio__isnull=True).exclude(numero_patrimonio__exact='')
        if self.value() == 'sem':
            return queryset.filter(numero_patrimonio__isnull=True) | queryset.filter(numero_patrimonio__exact='')
        return queryset


@admin.register(Equipamento)
class EquipamentoAdmin(admin.ModelAdmin):
    list_display = ('id', 'nome', 'quantidade_estoque', 'categoria', 'numero_patrimonio', 'status', 'funcionario', 'unidade')
    list_display_links = ('nome',)
    search_fields = ('nome', 'numero_patrimonio', 'categoria__nome', 'status')
    list_filter = ('categoria', 'status', 'funcionario', 'unidade', PatrimonioFilter)  # üëà filtro adicionado
    ordering = ('nome',)
    list_per_page = 10
    actions = ['exportar_pdf', 'exportar_excel']

    def exportar_pdf(self, request, queryset):
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="relatorio_equipamentos.pdf"'

        doc = SimpleDocTemplate(response, pagesize=A4)
        styles = getSampleStyleSheet()
        elements = []

        # Adiciona t√≠tulo
        titulo = Paragraph("Relat√≥rio de Equipamentos", styles['Title'])
        elements.append(titulo)
        elements.append(Spacer(1, 20))  # espa√ßo entre t√≠tulo e tabela

        # Cabe√ßalho da tabela
        data = [['Nome', 'Categoria', 'Status', 'Quantidade', 'Patrim√¥nio']]

        # Preenche as linhas da tabela
        for eq in queryset:
            data.append([
                eq.nome,
                str(eq.categoria),
                eq.status,
                str(eq.quantidade_estoque),
                eq.numero_patrimonio or 'Sem n√∫mero'
            ])

        # Cria a tabela
        table = Table(data, colWidths=[120, 100, 100, 70, 100])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.grey),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0,0), (-1,0), 8),
            ('GRID', (0,0), (-1,-1), 0.5, colors.black),
        ]))

        elements.append(table)
        doc.build(elements)
        return response

    exportar_pdf.short_description = "Exportar equipamentos selecionados para PDF"

    def exportar_excel(self, request, queryset):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Equipamentos"

        # Cabe√ßalho
        headers = ['Nome', 'Categoria', 'Status', 'Quantidade', 'Patrim√¥nio']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        # Preenche dados
        for row_num, eq in enumerate(queryset, 2):
            ws.cell(row=row_num, column=1, value=eq.nome)
            ws.cell(row=row_num, column=2, value=str(eq.categoria))
            ws.cell(row=row_num, column=3, value=eq.status)
            ws.cell(row=row_num, column=4, value=eq.quantidade_estoque)
            ws.cell(row=row_num, column=5, value=eq.numero_patrimonio or 'Sem n√∫mero')

        # Ajusta largura das colunas
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = length + 2

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=relatorio_equipamentos.xlsx'
        wb.save(response)
        return response

    exportar_excel.short_description = "Exportar equipamentos selecionados para Excel"


@admin.register(Movimentacao)
class MovimentacaoAdmin(admin.ModelAdmin):
    list_display = ('id', 'equipamento', 'data_movimentacao', 'origem_funcionario', 'destino_funcionario', 'origem_unidade', 'destino_unidade')
    search_fields = ('equipamento__nome',)
    list_filter = ('data_movimentacao', 'origem_unidade', 'destino_unidade', 'origem_funcionario', 'destino_funcionario')
    ordering = ('-data_movimentacao',)
    autocomplete_fields = ['equipamento', 'origem_funcionario', 'destino_funcionario', 'origem_unidade', 'destino_unidade']
    list_per_page = 15
    actions = ['exportar_pdf']

    def exportar_pdf(self, request, queryset):
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="relatorio_movimentacoes.pdf"'

        doc = SimpleDocTemplate(response, pagesize=A4)
        styles = getSampleStyleSheet()
        elements = []

        # T√≠tulo
        titulo = Paragraph("Relat√≥rio de Movimenta√ß√µes", styles['Title'])
        elements.append(titulo)
        elements.append(Spacer(1, 20))

        # Lista de movimenta√ß√µes
        for mov in queryset:
            texto = f"""
            <b>Equipamento:</b> {mov.equipamento}<br/>
            <b>Data:</b> {mov.data_movimentacao.strftime("%d/%m/%Y")}<br/>
            <b>Origem Func.:</b> {mov.origem_funcionario or '-'}<br/>
            <b>Destino Func.:</b> {mov.destino_funcionario or '-'}<br/>
            <b>Origem Unidade:</b> {mov.origem_unidade or '-'}<br/>
            <b>Destino Unidade:</b> {mov.destino_unidade or '-'}<br/>
            <b>Quantidade movimentada:</b> {mov.quantidade or '-'}<br/>
            <b>Estoque atual:</b> {mov.equipamento.quantidade_estoque}<br/>
            """
            elements.append(Paragraph(texto, styles['Normal']))
            elements.append(Spacer(1, 12))  # espa√ßo entre movimenta√ß√µes

        doc.build(elements)
        return response

    exportar_pdf.short_description = "Exportar movimenta√ß√µes selecionadas para PDF"
